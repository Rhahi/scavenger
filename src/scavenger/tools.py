import openpyxl as xl
import os
import csv
import datetime
import pathlib

def extract_horizontal(filepath, fields_list, sheetnames, range_condition=None):
    """Given a path to an excel file, extract the fields horizontally and return a list.

    Args:
        filepath (path): /path/to/file
        fields_list (list): ["list", "of", "field names", "to", "extract"]
        sheetnames (list): ["list", "of", "sheet", "names", "to", "try"], only the first success will be used.
        range_condition (string): string to limit the range, defaults to None

    Returns:
        dict: Matching results. If there was no matching cell or the match was empty, it will not be in the dict.
    """
    wb = xl.load_workbook(filepath, read_only=True)
    extracted = dict()
    
    # if there is a match, use that sheet. Otherwise, use the first sheet.
    for s in wb.sheetnames:
        if s in sheetnames:
            ws = wb[s]
            print("Using", s, end=" ")
            break
    else:
        ws = wb[wb.sheetnames[0]]
        print("Using", wb.sheetnames[0], end=" ")
    
    # define the search range of rows.
    if range_condition:
        row_range = find_range_with_condition(ws, range_condition)
    else:
        row_range = (1, ws.max_row+1)
    print(row_range)
    
    # extraction
    for r in range(row_range[0], row_range[1]):
        c = 1
        cat = None
        found = False
        while c <= ws.max_column:
            value = ws.cell(row=r, column=c).value
            # look for category
            if not found:
                cat = category(value, fields_list)
                if cat:
                    found = True
            # look for value in the category
            elif value is not None:
                if category(value, fields_list):
                    c-=1
                    found = False
                else:
                    extracted[cat] = value.strip()
                    found = False
            c+=1
    return extracted


def category(value, fields_list):
    """Given a category list of list, see if the value matches category or not.

    Args:
        value (str): value of a cell
        fields_list (list): [["categoryA1", "categoryA2"], ["categoryB]]

    Returns:
        str or None: str if it matches the category, None if there is no match.
    """
    if type(value) is str:
        v = value.replace(" ", "")
        for fl in fields_list:
            if v in fl:
                return v
    return None


def find_range_with_condition(ws, condition):
    """Given a worksheet, find a range of rows to use with matching condition.
    Only the first column is searched.

    Args:
        ws (worksheet): Openpyxl worksheet
        condition (str): match keyword

    Returns:
        tuple: (range_start:int, range_end:int)
    """
    range_start = 1
    range_end = ws.max_row
    i = 0
    # find left-most cell with matching value
    while i < ws.max_row:
        i+= 1
        if ws.cell(row=i, column=1).value == condition:
            range_start = i
            break
    # find left-most cell with un-matching, non-empty value
    while i < ws.max_row:
        i += 1
        if ws.cell(row=i, column=1).value is None:
            continue
        else:
            range_end = i
            break
    # above is the range.
    return (range_start, range_end)


def get_files_under(path, includes=None, excludes=None, oldest=None):
    """Given a path, return a list of files with matching conditions.

    Args:
        path (str): /path/to/directory
        includes (list, optional): list of str that the file name must include. Defaults to None.
        excludes (list, optional): list of str to exclude from the search. Defaults to None.
        oldest(int, optional): from and after year the file to retrieve. Defaults to None.
    """
    def check_in_list(filename, inlist):
        found_one=False
        found_all=True
        for i in inlist:
            if i in filename:
                found_one=True
            else:
                found_all=False
        return found_one, found_all

    result = []
    if not includes: includes=[]
    if not excludes: excludes=[]
    for root, _, files in os.walk(path, topdown=False):
        for filename in files:
            # guard against artifact from open file.
            if "$" in filename:
                continue
            exc, _ = check_in_list(filename, excludes)
            if exc:
                continue
            _, inc = check_in_list(filename, includes)
            if oldest:
                mtime = pathlib.Path(os.path.join(root, filename)).stat().st_mtime
                year = datetime.datetime.fromtimestamp(mtime).year
                if year < oldest:
                    continue
            if inc and filename.endswith(".xlsx"):
                result.append(os.path.join(root, filename))
    return result